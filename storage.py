import openpyxl
import hashlib
from datetime import datetime

FILE = "events.xlsx"

HEADERS = [
    "id", "event_name", "date", "venue",
    "city", "category", "url", "status", "last_seen"
]


def get_id(e):
    raw = e["event_name"] + e["date"] + e["venue"]
    return hashlib.md5(raw.encode()).hexdigest()


def load_sheet():
    try:
        wb = openpyxl.load_workbook(FILE)
        ws = wb.active
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    return wb, ws


def upsert_events(events):
    wb, ws = load_sheet()

    # store row OBJECT instead of values_only
    existing = {}

    for row in ws.iter_rows(min_row=2):
        existing[row[0].value] = row

    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    seen_ids = set()

    for e in events:
        eid = get_id(e)
        seen_ids.add(eid)

        row_data = [
            eid,
            e["event_name"],
            e["date"],
            e["venue"],
            e["city"],
            e["category"],
            e["url"],
            "Active",
            today
        ]

        # UPDATE existing
        if eid in existing:
            row = existing[eid]
            for i, val in enumerate(row_data):
                row[i].value = val

        # INSERT new
        else:
            ws.append(row_data)

    # mark expired
    mark_expired(ws, seen_ids)

    wb.save(FILE)
    print("Excel updated successfully ✅")


def mark_expired(ws, seen_ids):
    for row in ws.iter_rows(min_row=2):
        eid = row[0].value
        if eid not in seen_ids:
            row[7].value = "Expired"
